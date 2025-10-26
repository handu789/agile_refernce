import csv
import re
import os
import json
from openai import OpenAI
from docx import Document
from docx.shared import Pt, Inches, Cm
from docx.oxml.ns import qn
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.enum.section import WD_SECTION_START
from docx.shared import RGBColor
from docx.oxml import OxmlElement, ns
from datetime import datetime
from openpyxl import Workbook, load_workbook
from langchain.vectorstores import FAISS
from build_company_bank import LocalBGEEmbedding  # 假设你把之前的 LocalBertEmbedding 放在了一个模块里

client = OpenAI(api_key="<API_KEY>", base_url="<BASE_URL>")
embedding_model = LocalBGEEmbedding(model_name="BAAI/bge-large-zh", normalize=True)
vectordb = FAISS.load_local("company_index", embedding_model, allow_dangerous_deserialization=True)



def evaluate_and_rank_articles_multistep(dataset_paths, excel_path="rank_result.xlsx"):
    # ========== 初始化Excel文件 ==========
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["文章标题", "所属文件", "所属机构", "机构评分", "创新评分", "经济评分", "综合评分"])
        wb.save(excel_path)

    # ========== 遍历数据集 ==========
    for file_path in dataset_paths:
        print(f"正在处理文件：{file_path}")

        with open(file_path, "r", encoding="utf-8-sig") as csvfile:
            csv_reader = csv.reader(csvfile)

            while True:
                try:
                    headers = next(csv_reader)
                    data = next(csv_reader)
                except StopIteration:
                    break

                if len(headers) == 4 and len(data) == 4:
                    row = dict(zip(headers, data))
                    title = row['title']
                    passage = row['passage']

                    # ======= 第一步：识别所属机构 ========
                    identify_prompt = """你是一个科技新闻分析专家，请判断以下科技成果主要由哪一个机构提出或主导，并输出结构化信息（如有多个机构，请选择中重要的一个或者第一个出现的机构；如由高校与企业联合，也视为企业）。输出严格JSON格式如下，不得添加其他说明，绝对禁止使用任何形式的代码块标记（包括```json、```等）：

                    {
                      "机构名称": "xxx",
                      "所属机构": "企业"  // 或 "高校"
                    }
                    注意：绝对禁止使用任何形式的代码块标记（包括```json、```等）
                    新闻内容如下：
                    """ + passage

                    try:
                        identify_response = client.chat.completions.create(
                            model="YOUR_MODEL",
                            messages=[
                                {"role": "system", "content": "你是一个精确判断新闻所属机构的助手。"},
                                {"role": "user", "content": identify_prompt}
                            ],
                            temperature=0
                        )

                        org_json = json.loads(identify_response.choices[0].message.content.strip())
                        org_type = org_json.get("所属机构")
                        org_name = org_json.get("机构名称", "").strip()
                        print(f"✔️ {title} → 判断机构：{org_type} | 机构名称：{org_name}")

                    except Exception as e:
                        print(f"⚠️ 所属机构识别失败，跳过：{title} | 原因: {e}")
                        continue

                    # ======= 第二步：按机构类型处理打分 ========
                    matched_company, matched_level, score_value = None, None, 0.0

                    if org_type == "企业" and org_name:
                        retrieval_results = vectordb.similarity_search_with_score(org_name, k=1)
                        if retrieval_results:
                            doc, score = retrieval_results[0]
                            score_value = score
                            if score >= 0.85:
                                matched_company = doc.metadata.get("name")
                                matched_level = doc.metadata.get("level")
                                print(f"匹配到的企业:{matched_company}（相似度: {round(score, 4)}）")

                    # 构造评分用 Prompt
                    if org_type == "企业":
                        with open("company_score.txt", "r", encoding="utf-8") as f:
                            scoring_guide = f.read()
                        base_prompt = f"""你是一个严谨的企业类科技成果新闻评分助手，专门评估新闻中提及的【新发布或新成果】在以下三个维度的表现，并对每个维度进行 0-100 的打分：
                        新闻内容：{passage}
                        """

                        # 如果匹配到企业，才加入匹配信息段
                        if matched_company:
                            match_info = f"""
                        【向量匹配结果】
                        企业命中：{matched_company}
                        影响力等级：{matched_level or "无"}
                        相似度得分：{round(score_value, 4)}
                        """
                        else:
                            match_info = ""

                        scoring_prompt = base_prompt + match_info + f"""

                        评分说明：
                        {scoring_guide}

                        请输出如下 JSON 结构：
                        {{
                        "机构评分": 分数,
                        "创新评分": 分数,
                        "经济评分": 分数
                        }}
                        注意：
                        - 绝对禁止使用任何形式的代码块标记（包括```json、```等）。
                        - 分数必须是整数，范围在0-100之间。
                        """
                    else:  # 高校
                        with open("college_score.txt", "r", encoding="utf-8") as f:
                            scoring_guide = f.read()
                        scoring_prompt = f"""你是一个高校类科技成果新闻评分助手，专门评估新闻中提及的【新发布或新成果】在以下三个维度的表现，并对每个维度进行 0-100 的打分：
    新闻内容：{passage}

    评分说明：
    {scoring_guide}

    请输出如下 JSON 结构：
    {{
    "机构评分": 分数,
    "创新评分": 分数,
    "经济评分": 分数
    }}
    注意：
    - 绝对禁止使用任何形式的代码块标记（包括```json、```等）。
    - 分数必须是整数，范围在0-100之间。"""

                    # 模型打分
                    try:
                        score_response = client.chat.completions.create(
                            model="bot-20250419164220-f6p7d",
                            messages=[
                                {"role": "system", "content": "你是一个科技新闻评分助手。"},
                                {"role": "user", "content": scoring_prompt}
                            ],
                            temperature=0
                        )
                        score_data = json.loads(score_response.choices[0].message.content.strip())

                        company_score = score_data["机构评分"]
                        innovation_score = score_data["创新评分"]
                        economic_score = score_data["经济评分"]

                        # 按类型赋权
                        if org_type == "公司":
                            total_score = round(0.1 * company_score + 0.6 * innovation_score + 0.3 * economic_score, 2)
                        else:
                            total_score = round(0.2 * company_score + 0.7 * innovation_score + 0.1 * economic_score, 2)

                        # 写入 Excel
                        wb = load_workbook(excel_path)
                        ws = wb.active
                        ws.append([
                            title,
                            os.path.basename(file_path),
                            org_type,
                            company_score,
                            innovation_score,
                            economic_score,
                            total_score
                        ])
                        wb.save(excel_path)

                        print(f"✅ 写入完成 | 标题: {title} | 文件: {os.path.basename(file_path)} | 机构: {org_type} | 机构评分: {company_score} | 创新评分: {innovation_score} | 经济评分: {economic_score} | 综合评分: {total_score}")


                    except Exception as e:
                        print(f"⚠️ 打分失败，跳过：{title} | 错误：{e}")

    # ========== 排序Excel ==========
    print("\n所有文件处理完成，开始排序。")
    wb = load_workbook(excel_path)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    header, data_rows = all_rows[0], all_rows[1:]
    sorted_rows = sorted(data_rows, key=lambda x: x[-1], reverse=True)
    ws.delete_rows(2, ws.max_row)
    for row in sorted_rows:
        ws.append(row)
    wb.save(excel_path)
    print(f"✅ 排序完成，结果已保存至 {excel_path}")







# ====== 你自己手动指定要处理的CSV文件路径 ======
dataset_path = [
    '../20250602csv/36ke.csv',
    '../20250602csv/dongdiankeji.csv',
    '../20250602csv/jikegongyuan.csv',
    '../20250602csv/jiqizhixin.csv',
    '../20250602csv/jizhijulebu.csv',
    '../20250602csv/liangziwei.csv'
    #'../20250602csv/xinzhiyuan.csv'
]
evaluate_and_rank_articles_multistep(dataset_path)