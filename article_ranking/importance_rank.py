import csv
import re
import os
import json
from openai import OpenAI
from docx import Document
from docx.shared import Pt, Inches, Cm
from docx.oxml.ns import qn
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.enum.section import WD_SECTION_START
from docx.shared import RGBColor
from docx.oxml import OxmlElement, ns
from datetime import datetime
from openpyxl import Workbook, load_workbook

client = OpenAI(api_key="<API_KEY>", base_url="<BASE_URL>")

def evaluate_and_rank_articles(dataset_paths, excel_path="rank_result.xlsx", scoring_guide_path="scoring_criteria.txt"):
    # ====== 准备 Excel 文件 ======
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["文章标题", "所属文件", "企业评分", "创新评分", "经济评分", "综合评分"])
        wb.save(excel_path)

    # ====== 读取评分标准文件 ======
    with open(scoring_guide_path, "r", encoding="utf-8") as f:
        scoring_guide = f.read()

    # ====== 构造系统提示词 ======
    system_prompt = f"""你是一个严谨的科技新闻评分助手，专门评估新闻中提及的【新发布或新成果】在以下三个维度的表现，并对每个维度进行 0-100 的打分。

维度包括：
1. 企业影响力
2. 技术创新程度
3. 经济效益

评估要求：
- 在评分过程中，如遇到不确定的企业、技术或经济效益，请主动联网搜索查询相关资料，基于最新权威信息进行判断，不得仅凭已有知识推测。
- 仅对【新闻中明确发布的科技成果、新产品、新技术或明确展示demo的项目】进行打分。
- 如果新闻**主要是评论、回顾、批评、预测未来趋势**，而**没有明确新成果/新发布/新应用**，则对应各维度酌情降低评分，避免高分。
- 综合语义理解、关键词匹配和联网查询结果，作出严谨、客观的打分。
- 所有评分标准依据以下内容：

{scoring_guide}

你的输出格式应为严格的 JSON 格式，格式如下：

{{
"企业评分": 分数,
"创新评分": 分数,
"经济评分": 分数
}}

注意：
- 绝对禁止使用任何形式的代码块标记（包括```json、```等）。
- 分数必须是整数，范围在0-100之间。
"""

    # ====== 遍历数据集文件列表 ======
    for file_path in dataset_paths:
        print(f"正在处理文件：{file_path}")

        with open(file_path, "r", encoding="utf-8") as csvfile:
            csv_reader = csv.reader(csvfile)

            while True:
                try:
                    headers = next(csv_reader)
                    data = next(csv_reader)
                except StopIteration:
                    break

                if len(headers) == 4 and len(data) == 4:
                    row = dict(zip(headers, data))

                    # 🔹调用大模型评分
                    check_response = client.chat.completions.create(
                        model="YOUR_MODEL",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {
                                "role": "user",
                                "content": f"""请对以下新闻进行打分：
内容如下：
{row['passage']}"""
                            }
                        ],
                        temperature=0.0
                    )

                    result_text = check_response.choices[0].message.content.strip()
                    try:
                        score_data = json.loads(result_text)
                        company_score = score_data["企业评分"]
                        innovation_score = score_data["创新评分"]
                        economic_score = score_data["经济评分"]

                        total_score = round(
                            0.3 * company_score +
                            0.5 * innovation_score +
                            0.2 * economic_score, 2
                        )

                        # 写入Excel
                        wb = load_workbook(excel_path)
                        ws = wb.active
                        ws.append([
                            row["title"],
                            os.path.basename(file_path),
                            company_score,
                            innovation_score,
                            economic_score,
                            total_score
                        ])
                        wb.save(excel_path)

                        print(f"✅ 写入：{row['title']} - 企业: {company_score}, 创新: {innovation_score}, 经济: {economic_score}, 综合: {total_score}")
                        if hasattr(check_response, "references"):
                            print(check_response.references)

                    except json.JSONDecodeError:
                        print(f"⚠️ JSON解析失败，跳过：{row['title']}")

    # ====== 排序Excel（按综合评分降序） ======
    print(f"\n所有文件处理完成，开始排序。")

    wb = load_workbook(excel_path)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    data_rows = all_rows[1:]

    data_rows_sorted = sorted(data_rows, key=lambda x: x[-1], reverse=True)

    ws.delete_rows(2, ws.max_row)
    for row in data_rows_sorted:
        ws.append(row)

    wb.save(excel_path)
    print(f"✅ 排序完成，结果已保存至 {excel_path}")


# ====== 你自己手动指定要处理的CSV文件路径 ======
dataset_path = [
    '../csv_folder/36ke.csv',
    '../csv_folder/dongdiankeji.csv',
    '../csv_folder/jikegongyuan.csv',
    '../csv_folder/jiqizhixin.csv',
    '../csv_folder/jizhijulebu.csv',
    '../csv_folder/liangziwei.csv',
    '../csv_folder/xinzhiyuan.csv',
]
evaluate_and_rank_articles(dataset_path)